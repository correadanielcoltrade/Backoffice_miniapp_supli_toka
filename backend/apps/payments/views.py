from io import BytesIO

import openpyxl
from django.conf import settings
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from drf_spectacular.utils import extend_schema
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.inventory.services import InsufficientStockError, deduct_for_paid_order
from apps.orders.models import Order
from apps.users.models import Role
from apps.users.permissions import HasRole

from .filters import PaymentFilter
from .models import PaymentTransaction
from .serializers import (
    PaymentTransactionSerializer,
    TokaPaymentWebhookSerializer,
)


class PaymentTransactionViewSet(viewsets.ReadOnlyModelViewSet):
    """Modulo de Seguimiento de Transacciones de Pagos (solo lectura en el back office)."""

    queryset = PaymentTransaction.objects.select_related("order").all()
    serializer_class = PaymentTransactionSerializer
    permission_classes = [HasRole]
    allowed_roles = [Role.SALES, Role.LOGISTIC]
    filterset_class = PaymentFilter
    search_fields = ["payment_number", "customer_name", "toka_customer_id"]
    ordering_fields = ["created_at", "amount", "status"]

    @action(detail=False, methods=["get"])
    def export(self, request):
        """
        Descarga en Excel (.xlsx) los pagos que cumplan los filtros aplicados
        (mismos parametros que el listado: created_after, created_before, status,
        toka_customer_id, search). Incluye todos los detalles del pago.
        """
        queryset = self.filter_queryset(self.get_queryset()).order_by("-created_at")

        headers = [
            "Fecha de creacion",
            "Id Cliente",
            "Nombre Cliente",
            "# Pago",
            "PaymentId Toka",
            "Monto",
            "Estado de pago",
            "Pedido",
            "Fecha de confirmacion",
        ]
        money_col = 6  # Monto

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pagos"

        header_fill = PatternFill("solid", fgColor="1F3B57")
        header_font = Font(bold=True, color="FFFFFF")
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        def local(dt):
            return timezone.localtime(dt).strftime("%Y-%m-%d %H:%M") if dt else ""

        for p in queryset:
            order = p.order
            # Nombre y numero de pedido igual que en la tabla (via el pedido).
            name = (order.recipient_name if order and order.recipient_name
                    else p.customer_name)
            order_number = order.order_number if order else ""
            ws.append([
                local(p.created_at),
                p.toka_customer_id,
                name,
                p.payment_number,
                p.provider_payment_id,
                float(p.amount) if p.amount is not None else None,
                p.get_status_display(),
                order_number,
                local(p.confirmed_at),
            ])

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=money_col).number_format = '"$"#,##0.00'

        widths = [18, 20, 24, 26, 26, 14, 18, 14, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="pagos_{stamp}.xlsx"'
        )
        return response


class TokaPaymentWebhookView(APIView):
    """
    Webhook llamado por el BACKEND DE TOKA para confirmar un pago.
    Al confirmarse:
      1. Se registra/actualiza la transaccion de pago.
      2. Si hay pedido asociado, se marca como PAGADO.
      3. Se descuenta el inventario de los productos del pedido (una sola vez).

    Autenticacion: header 'X-Toka-Signature' == TOKA_WEBHOOK_SECRET.
    """

    authentication_classes = []
    permission_classes = [AllowAny]

    @extend_schema(
        request=TokaPaymentWebhookSerializer,
        responses=PaymentTransactionSerializer,
        description="Webhook llamado por el backend de Toka para confirmar un pago.",
    )
    def post(self, request):
        signature = request.headers.get("X-Toka-Signature", "")
        if not settings.TOKA_WEBHOOK_SECRET or signature != settings.TOKA_WEBHOOK_SECRET:
            return Response(
                {"detail": "Firma del webhook invalida."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        serializer = TokaPaymentWebhookSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        with transaction.atomic():
            payment, _ = PaymentTransaction.objects.select_for_update().update_or_create(
                payment_number=data["payment_number"],
                defaults={
                    "toka_customer_id": data["toka_customer_id"],
                    "customer_name": data["customer_name"],
                    "amount": data.get("amount"),
                    "status": data["status"],
                    "raw_payload": request.data,
                },
            )

            order = None
            order_id = data.get("order_id")
            if order_id:
                order = Order.objects.select_for_update().filter(pk=order_id).first()
                if order:
                    payment.order = order

            confirmed = data["status"] == PaymentTransaction.Status.CONFIRMED
            if confirmed:
                payment.confirmed_at = timezone.now()

            payment.save()

            # Descontar inventario solo una vez cuando el pago es confirmado
            if confirmed and order and not order.stock_deducted:
                try:
                    deduct_for_paid_order(order, reference=payment.payment_number)
                except InsufficientStockError as exc:
                    return Response(
                        {"detail": str(exc)},
                        status=status.HTTP_409_CONFLICT,
                    )
                order.status = Order.Status.PAID
                order.stock_deducted = True
                order.save(update_fields=["status", "stock_deducted", "updated_at"])

        return Response(
            PaymentTransactionSerializer(payment).data,
            status=status.HTTP_200_OK,
        )
