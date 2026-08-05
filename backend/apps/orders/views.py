from io import BytesIO

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import mixins
from rest_framework import status as http_status
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.users.models import Role
from apps.users.permissions import HasRole

from .filters import OrderFilter
from .models import Customer, CustomerAddress, Order, TrackingEvent
from .serializers import (
    CustomerAddressSerializer,
    CustomerSerializer,
    OrderSerializer,
    TrackingUpdateSerializer,
)


class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.prefetch_related("addresses").all()
    serializer_class = CustomerSerializer
    permission_classes = [HasRole]
    allowed_roles = [Role.SALES, Role.LOGISTIC]
    search_fields = ["full_name", "toka_customer_id", "contact_number", "email"]
    ordering_fields = ["full_name", "created_at"]


class CustomerAddressViewSet(viewsets.ModelViewSet):
    """Direcciones guardadas por cliente (tabla independiente por cliente)."""

    queryset = CustomerAddress.objects.select_related("customer").all()
    serializer_class = CustomerAddressSerializer
    permission_classes = [HasRole]
    allowed_roles = [Role.SALES, Role.LOGISTIC]
    filterset_fields = ["customer", "is_default"]
    search_fields = ["complete_address", "suburb", "municipality"]


class OrderViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """
    Modulo de Gestion de Pedidos. Los pedidos SOLO se crean desde la mini app
    (flujo /v1); no se permite la creacion manual por API (sin POST a la lista).
    """

    queryset = (
        Order.objects.select_related("customer", "saved_address")
        .prefetch_related("items__product", "tracking_events__created_by")
        .all()
    )
    serializer_class = OrderSerializer
    permission_classes = [HasRole]
    allowed_roles = [Role.SALES, Role.LOGISTIC]
    filterset_class = OrderFilter
    search_fields = [
        "recipient_name", "full_address", "customer__full_name", "tracking_guide",
    ]
    ordering_fields = ["created_at", "total_amount", "status"]

    @action(detail=False, methods=["get"])
    def export(self, request):
        """
        Descarga en Excel (.xlsx) los pedidos que cumplan los filtros aplicados
        (mismos parametros que el listado: created_after, created_before,
        status, customer, search, etc.).

        Formato picking: UNA FILA POR PRODUCTO. Los datos del pedido se repiten
        en cada fila de sus productos, para poder filtrar/sumar por SKU y
        despachar comodamente en Excel.
        """
        queryset = self.filter_queryset(self.get_queryset())

        headers = [
            "N° Pedido",
            "Cliente",
            "Nombre Completo",
            "Numero de contacto",
            "Direccion Completa",
            "Complemento de direccion",
            "Colonia",
            "Ciudad / Alcaldia",
            "Estado",
            "Codigo postal",
            "Producto",
            "SKU",
            "Cantidad",
            "Precio unitario",
            "Subtotal",
            "Total del pedido",
            "Estado del pedido",
            "Fecha de creacion",
        ]
        # Columnas con formato de moneda (1-indexado, para openpyxl).
        money_cols = {14, 15, 16}  # Precio unitario, Subtotal, Total del pedido
        qty_col = 13

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedidos"

        header_fill = PatternFill("solid", fgColor="1F3B57")
        header_font = Font(bold=True, color="FFFFFF")
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        def order_base(o):
            return [
                o.order_number,
                o.customer.full_name,
                o.recipient_name,
                o.contact_number,
                o.full_address,
                o.address_complement,
                o.colonia,
                o.city_alcaldia,
                o.state,
                o.postal_code,
            ]

        def order_tail(o):
            return [
                float(o.total_amount),
                o.get_status_display(),
                timezone.localtime(o.created_at).strftime("%Y-%m-%d %H:%M"),
            ]

        for o in queryset:
            items = list(o.items.all())
            if items:
                for it in items:
                    ws.append(
                        order_base(o)
                        + [
                            it.product.description,
                            it.product.sku,
                            it.quantity,
                            float(it.unit_price),
                            float(it.subtotal),
                        ]
                        + order_tail(o)
                    )
            else:
                # Pedido sin productos: una fila con las columnas de producto vacias.
                ws.append(order_base(o) + ["", "", "", "", ""] + order_tail(o))

        # Formato de moneda y cantidad para las filas de datos.
        for row in range(2, ws.max_row + 1):
            for col in money_cols:
                ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=qty_col).alignment = Alignment(
                horizontal="center"
            )

        # Anchos de columna aproximados para que se lea bien de una.
        widths = [14, 22, 22, 16, 30, 22, 18, 20, 16, 12, 28, 16, 10, 15, 13, 15, 18, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Fila de encabezados fija + autofiltro para ordenar/filtrar en Excel.
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="pedidos_{stamp}.xlsx"'
        )
        return response

    @action(detail=True, methods=["post"])
    def tracking(self, request, pk=None):
        """
        POST /api/orders/{id}/tracking/  — registra un paso de la entrega.

        Body: {status, note?, tracking_guide?, carrier?}. Actualiza el estado
        ACTUAL del pedido, guarda guia/transportadora si vienen, y agrega un
        evento al historial (append-only). Al llegar a DELIVERED sincroniza el
        estado del pedido a 'Entregado'.
        """
        order = self.get_object()
        ser = TrackingUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        old_guide = order.tracking_guide
        new_guide = (data.get("tracking_guide") or "").strip()
        guide_changed = bool(new_guide) and new_guide != old_guide

        update_fields = ["tracking_status", "updated_at"]
        order.tracking_status = data["status"]
        # Guia/transportadora "sticky": solo se actualizan si viene un valor no
        # vacio, para que avanzar el estado no borre datos ya cargados.
        if new_guide:
            order.tracking_guide = new_guide
            update_fields.append("tracking_guide")
        if data.get("carrier"):
            order.carrier = data["carrier"]
            update_fields.append("carrier")
        # Sincroniza el estado del pedido cuando la entrega se completa.
        if data["status"] == Order.DeliveryStatus.DELIVERED:
            order.status = Order.Status.DELIVERED
            update_fields.append("status")
        order.save(update_fields=update_fields)

        # Deja constancia del cambio de guia en el historial (util para un
        # re-despacho tras una novedad): "Guia actualizada: anterior -> nueva".
        note_parts = []
        if guide_changed:
            if old_guide:
                note_parts.append(f"Guía actualizada: {old_guide} → {new_guide}")
            else:
                note_parts.append(f"Guía asignada: {new_guide}")
        if data.get("note"):
            note_parts.append(data["note"].strip())

        TrackingEvent.objects.create(
            order=order,
            status=data["status"],
            note=" — ".join(note_parts),
            created_by=request.user if request.user.is_authenticated else None,
        )

        order = self.get_queryset().get(pk=order.pk)  # recarga con prefetch
        return Response(
            OrderSerializer(order, context={"request": request}).data,
            status=http_status.HTTP_200_OK,
        )
